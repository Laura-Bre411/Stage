"""
Procédure pour calculer le spectre de réponse en déplacement :

a) Obtenir la transformée de Fourier de l'accélérogramme (domaine fréquentiel).

b) Appliquer la fonction de transfert en déplacemen et modifier la réponse en fonction de la fonction de transfert de l’O.H.

c) Revenir dans le domaine temporel en effectuant la transformée de Fourier inverse.

d) Répétez les étapes pour une gamme variée de fréquences correspondant à différents oscillateurs ( et identifier le déplacement
maximal pour obtenir le spectre de réponse en déplacement.

Auteur : Laura Brémont

Dernière version : 30/01/26
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from scipy.fft import fft, fftfreq


###Fréquences de l'O.H à tester
frequencies = np.logspace(-1, 1.5, num=150)


def accelero_to_spectre(frequencies, accel_data, dt, amort):
    """
    Calcule les spectres de réponse en déplacement, vitesse et accélération pour un accélérogramme donné.

    Args:
        frequencies (array): Fréquences propres des oscillateurs harmoniques (en Hz).
        accel_data (array): Données d'accélération (accélérogramme) en m/s^2 ou g.
        dt (float): Intervalle de temps entre les échantillons du signal (en secondes).
        amort (float): Coefficient d'amortissement (par exemple, 0.05 pour 5%).

    Returns:
        dict: Contient les spectres de déplacement, vitesse et accélération correspondant aux fréquences propres.
    """

    # Nombre de points et fréquence d'échantillonnage
    n_points = len(accel_data)
    sampling_freq = 1 / dt

    # Transformée de Fourier de l'accélérogramme
    fft_accel = fft(accel_data)
    frequencies_fft = fftfreq(n_points, dt)    # Fréquences associées (en Hz)
    print(frequencies_fft)
    # Conserver seulement les fréquences positives
    #correct_freq_indices = np.where(frequencies_fft >= 0 and frequencies_fft < sampling_freq/2, 1, 0)  #si la condition est vérifiée, l'indice
    fft_accel = np.array([frequencies_fft[i] for i in range(n_points) if frequencies_fft[i] > 0 and frequencies_fft[i] < sampling_freq/2])
    frequencies_fft = np.array([frequencies_fft[i] for i in range(n_points) if frequencies_fft[i] > 0 and frequencies_fft[i] < sampling_freq/2])

    # Calcul des spectres pour les fréquences propres spécifiées
    spectre_d = []  # Déplacement
    spectre_v = []  # Pseudo-vitesse
    spectre_a = []  # Pseudo-accélération

    for fn in frequencies:  # Boucle sur les fréquences propres des oscillateurs harmoniques
        omega_n = 2 * np.pi * fn  # Pulsation propre (rad/s)

        # Fonction de transfert H(omega) pour le déplacement
        h_omega = 1 / np.sqrt((1 - (frequencies_fft / omega_n) ** 2) ** 2 + (2 * amort * frequencies_fft / omega_n) ** 2)

        # Résolution dans le domaine fréquentiel
        fft_displacement_response = fft_accel * h_omega / (-(frequencies_fft ** 2))

        # Transformée inverse pour obtenir la réponse dans le domaine temporel
        displacement_response_time = np.abs(fft_displacement_response)

        # Obtenir le déplacement maximal
        max_displacement = np.max(displacement_response_time)
        spectre_d.append(max_displacement)

        # Calcul des pseudo-grandeurs :
        max_pseudo_velocity = max_displacement * omega_n
        max_pseudo_acceleration = max_displacement * omega_n ** 2

        spectre_v.append(max_pseudo_velocity)
        spectre_a.append(max_pseudo_acceleration)

    # Conversion des résultats en numpy arrays pour un traitement plus simple
    freqs = np.array(frequencies)
    spectre_d = np.array(spectre_d)
    spectre_v = np.array(spectre_v)
    spectre_a = np.array(spectre_a)
    
    return {
        "Frequences": freqs,
        "Deplacement": spectre_d,
        "Pseudo_Vitesse": spectre_v,
        "Pseudo_Acceleration": spectre_a
    }


def affichage(temps, accel, resultats):
    """ 
    Affiche les graphiques des différentes réponses et annote les maxima. 
    Entrées : 
        - temps : array
            Vecteur temps pour le signal d'entrée
        - accel : array
            Accélérogramme (signal d'entrée)
        - resultats : dict
            Résultats de la méthode SRO NIGAM comprenant les spectres de déplacement,
            pseudo-vitesse et pseudo-accélération
    """
    
    fig, axs = plt.subplots(2, 2, figsize=(12, 8))
    print(resultats)
    # --- Subplot 1: Accélération ---
    axs[0, 0].grid(True, which="both", ls="-")
    axs[0, 0].plot(temps, accel)
    axs[0, 0].set_xlabel("Temps (s)")
    axs[0, 0].set_ylabel("Accélération (m/s² ou g)")
    axs[0, 0].set_title("Perturbation en accélération")
    
    # --- Subplot 2: Spectre de déplacement relatif ---
    axs[0, 1].grid(True, which="both", ls="-")
    axs[0, 1].loglog(resultats['Frequences'], resultats['Deplacement'], color='purple', label='Amortissement 5%')
    axs[0, 1].set_xlabel("Fréquence de l'O.H (Hz)")
    axs[0, 1].set_ylabel("Déplacement (m)")
    axs[0, 1].set_title("Spectre de déplacement relatif")
    axs[0, 1].legend(loc='best')
    
    # Repérer et annoter le maximum pour le déplacement
    max_disp_idx = np.argmax(resultats['Deplacement'])
    max_disp_val = resultats['Deplacement'][max_disp_idx]
    max_disp_freq = resultats['Frequences'][max_disp_idx]
    axs[0, 1].annotate(f'Max: {max_disp_val:.2e} m\nFreq: {max_disp_freq:.2f} Hz',
                       xy=(max_disp_freq, max_disp_val),
                       xytext=(-40, -90),
                       textcoords='offset points',
                       arrowprops=dict(facecolor='black', shrink=0.05),
                       fontsize=10,
                       bbox=dict(boxstyle="round",pad=0.3, edgecolor="black", facecolor="white", alpha=0.5))

    # --- Subplot 3: Spectre de pseudo-vitesse ---
    axs[1, 0].grid(True, which="both", ls="-")
    axs[1, 0].loglog(resultats['Frequences'], resultats['Pseudo_Vitesse'], color='green', label='Amortissement 5%')
    axs[1, 0].set_xlabel("Fréquence de l'O.H (Hz)")
    axs[1, 0].set_ylabel("Pseudo-vitesse (m/s)")
    axs[1, 0].set_title("Spectre de pseudo-vitesse")
    axs[1, 0].legend(loc='best')
    
    # Repérer et annoter le maximum pour la vitesse
    max_vel_idx = np.argmax(resultats['Pseudo_Vitesse'])
    max_vel_val = resultats['Pseudo_Vitesse'][max_vel_idx]
    max_vel_freq = resultats['Frequences'][max_vel_idx]
    axs[1, 0].annotate(f'Max: {max_vel_val:.2e} m/s\nFreq: {max_vel_freq:.2f} Hz',
                       xy=(max_vel_freq, max_vel_val),
                       xytext=(-40, -90),
                       textcoords='offset points',
                       arrowprops=dict(facecolor='black', shrink=0.05),
                       fontsize=10,
                       bbox=dict(boxstyle="round",pad=0.3, edgecolor="black", facecolor="white", alpha=0.5))
    
    # --- Subplot 4: Spectre de pseudo-accélération ---
    axs[1, 1].grid(True, which="both", ls="-")
    axs[1, 1].loglog(resultats['Frequences'], resultats['Pseudo_Acceleration'], color='coral', label='Amortissement 5%')
    axs[1, 1].set_xlabel("Fréquence de l'O.H (Hz)")
    axs[1, 1].set_ylabel("Pseudo-accélération (m/s² ou g)")
    axs[1, 1].set_title("Spectre de pseudo-accélération")
    axs[1, 1].legend(loc='best')
    
    # Repérer et annoter le maximum pour l'accélération
    max_acc_idx = np.argmax(resultats['Pseudo_Acceleration'])
    max_acc_val = resultats['Pseudo_Acceleration'][max_acc_idx]
    max_acc_freq = resultats['Frequences'][max_acc_idx]
    axs[1, 1].annotate(f'Max: {max_acc_val:.2e} m/s²\nFreq: {max_acc_freq:.2f} Hz',
                       xy=(max_acc_freq, max_acc_val),
                       xytext=(-40,-90),
                       textcoords='offset points',
                       arrowprops=dict(facecolor='black', shrink=0.05),
                       fontsize=10,
                       bbox=dict(boxstyle="round", pad=0.3, edgecolor="black", facecolor="white", alpha=0.5))
    
    plt.tight_layout()
    plt.show()
    
    
# Lire le fichier Excel (feuille Calculs)
df_calculs = pd.read_excel("Spectre.xlsx", sheet_name="Calculs", header=None, engine="openpyxl")

dt = df_calculs.iloc[10, 3]  # Cellule D11 (index 0 donc ligne 10, colonne 3)
amort = df_calculs.iloc[19, 3]  # Cellule D20 (index 19, colonne 3)
print(dt, amort)

# Lire l'accélérogramme (feuille A)
df_acc = pd.read_excel("Spectre.xlsx", sheet_name="A", engine="openpyxl")
accel_data = df_acc.iloc[2:, 1].values  # Colonne 2 (les données commencent à la ligne 3)
temps = df_acc.iloc[2:, 0].values  # Colonne 1 (sans la ligne de titre)
print(accel_data)
print(temps)

# Lancer le calcul (fonction supposée existante)
res = accelero_to_spectre(frequencies, accel_data, dt, amort)  # Résultat attendu sous la forme d'une liste ou d'un tableau

# Affichage des données (si nécessaire pour vérification)
affichage(temps, accel_data, res)

# Charger le fichier existant avec openpyxl pour modification sans écrasement

try:
    # Charger le fichier Excel
    wb = load_workbook("Spectre.xlsx")
    
    # Créer/Réinitialiser la feuille "RES"
    if "RES" in wb.sheetnames:
        ws = wb["RES"]
        ws.delete_rows(1, ws.max_row)  # Vider les anciennes données
    else:
        ws = wb.create_sheet(title="RES")
    
    # Ajouter les en-têtes
    headers = ["Fréq. propres", "Déplacement rel.", "Ps. Vitesse", "Ps. Accélération"]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    
    # Ajouter les données
    for row_idx, row in enumerate(res.values(), start=1):  # Commence à la ligne 2
        for col_idx, value in enumerate(row, start=2):
            ws.cell(row=col_idx, column=row_idx, value=value)   #avoir 4 colonnes (feuille RES) à partir de 4 lignes (dict res)
    
    # Sauvegarder les modifications
    wb.save("Spectre.xlsx")
    wb.close() 

      
     
except Exception as e:
    print(f"Une erreur s'est produite : {e}")
    
    
    
"""
#Exemple simple pour comprendre le principe

# Number of sample points
N = 600
# sample spacing
T = 1.0 / 800.0
x = np.linspace(0.0, N*T, N, endpoint=False)
y = np.sin(50.0 * 2.0*np.pi*x) + 0.5*np.sin(80.0 * 2.0*np.pi*x)
yf = fft(y)
xf = fftfreq(N, T)[:N//2]
import matplotlib.pyplot as plt
plt.plot(xf, 2.0/N * np.abs(yf[0:N//2]))
plt.grid()
plt.show()
"""
