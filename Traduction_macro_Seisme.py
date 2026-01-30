"""Composition du programme :
-   Fonction sro_nigam
-   Partie lecture d'Excel, appel, et écriture des résultats

Auteur : Laura Brémont
Dernière version : 29/01/26
"""

import pandas as pd
from openpyxl import load_workbook
# Pour conserver les macros, utiliser openpyxl ou xlwings 
# uniquement, sans pandas pour écrire les données directement
import numpy as np
import matplotlib.pyplot as plt
from matplotlib import rcParams
rcParams['font.family'] = 'serif'
rcParams['font.size'] = 14


###Fréquences de l'O.H à tester
frequencies = np.logspace(-1, 1.5, num=150)


def sro_nigam(frequencies, accelerogram, dt, damping):
    """
    Calcule le Spectre de Réponse d'Oscillateur (SRO) selon la méthode de NIGAM & JENNINGS.
    
    Entrées :
    frequencies : array
        Tableau des fréquences propres (Hz) pour lesquelles calculer le spectre
    accelerogram : array
        Signal d'accélération (accélérogramme)
    dt : float
        Pas de temps du signal (s)
    damping : float
        Amortissement réduit xi
        

    Sortie : dict
        Dictionnaire contenant :
        - 'D': Spectre de déplacement relatif
        - 'V': Spectre de pseudo-vitesse
        - 'A': Spectre de pseudo-accélération
    """
    
    # Conversion en tableaux numpy
    freqs = np.asarray(frequencies)
    acc = np.asarray(accelerogram)
    
    # Initialisation des résultats
    n_freqs = len(freqs)
    spectre_d = np.zeros(n_freqs)
    spectre_v = np.zeros(n_freqs)
    spectre_a = np.zeros(n_freqs)
    
    # Pré-calculs constants
    xi = damping
    xi_sq = np.sqrt(1 - xi**2)  # Amortissement réduit (sqrt(1-xi^2))
    
    # Boucle sur chaque fréquence (comme dans la macro VBA "For fn = 1 To Nf_")
    for i, f in enumerate(freqs):
        omega = 2 * np.pi * f
        omega_d = omega * xi_sq
        w2 = omega**2
        w3 = omega**3
        
        # --- Calcul des matrices A et B (Méthode NIGAM) ---
        # Voir doc R5.05.01 Page 5 
        
        exp_val = np.exp(-xi * omega * dt)
        sin_val = np.sin(omega_d * dt)
        cos_val = np.cos(omega_d * dt)
        
        # Coefficients de la matrice A
        a11 = exp_val * (xi / xi_sq * sin_val + cos_val)
        a12 = exp_val / omega_d * sin_val
        a21 = -omega / xi_sq * exp_val * sin_val
        a22 = exp_val * (cos_val - xi / xi_sq * sin_val)
        
        # Matrice A
        A_mat = np.array([[a11, a12],
                          [a21, a22]])
        
        # Coefficients intermédiaires pour la matrice B
        c1 = (2 * xi**2 - 1) / (w2 * dt)
        c2 = (2 * xi) / (w3 * dt)
        
        # Coefficients de la matrice B (partie liée au chargement)
        # Note: termes identiques à la macro VBA
        b11 = exp_val * ((c1 + xi/omega) * sin_val/omega_d + (c2 + 1/w2) * cos_val) - c2
        b12 = -exp_val * (c1 * sin_val/omega_d + c2 * cos_val) - 1/w2 + c2
        
        b21 = exp_val * ((c1 + xi/omega) * (cos_val - xi/xi_sq * sin_val) - 
                         (c2 + 1/w2) * (omega_d * sin_val + xi*omega * cos_val)) + 1/(w2 * dt)
                         
        b22 = -exp_val * (c1 * (cos_val - xi/xi_sq * sin_val) - 
                          c2 * (omega_d * sin_val + xi*omega * cos_val)) - 1/(w2 * dt)
        
        B_mat = np.array([[b11, b12],
                          [b21, b22]])
        
        # --- Résolution temporelle (Boucle sur le temps) ---
        # Initialisation vecteur état Q = [déplacement, vitesse]
        q = np.zeros(2)
        d_max = 0.0
        
        # On parcourt l'accélérogramme pas à pas
        # Correspond à "For tn = 2 To NB_val_T_"
        for t in range(1, len(acc)):
            alpha_vec = np.array([acc[t-1], acc[t]]) # [alpha(t-dt), alpha(t)]
            
            # Relation de récurrence : Q(t) = A * Q(t-dt) + B * Alpha
            q = A_mat @ q + B_mat @ alpha_vec
            
            # Recherche du maximum absolu du déplacement
            if abs(q[0]) > d_max:
                d_max = abs(q[0])
        
        # Stockage des résultats pour cette fréquence
        spectre_d[i] = d_max                  # Déplacement relatif max
        spectre_v[i] = omega * d_max          # Pseudo-vitesse (omega * SD)
        spectre_a[i] = w2 * d_max             # Pseudo-accélération (omega^2 * SD)

    return {
        "Frequences": freqs,
        "Deplacement": spectre_d,
        "Pseudo_Vitesse": spectre_v,
        "Pseudo_Acceleration": spectre_a
    }

def affichage(temps, accel, resultats):
    """ 
    Affiche les graphiques des différentes réponses et annote les maxima. 
    Entrées : 
        - temps : array
            Vecteur temps pour le signal d'entrée
        - accel : array
            Accélérogramme (signal d'entrée)
        - resultats : dict
            Résultats de la méthode SRO NIGAM comprenant les spectres de déplacement,
            pseudo-vitesse et pseudo-accélération
    """
    
    fig, axs = plt.subplots(2, 2, figsize=(12, 8))
    print(resultats)
    # --- Subplot 1: Accélération ---
    axs[0, 0].grid(True, which="both", ls="-")
    axs[0, 0].plot(temps, accel)
    axs[0, 0].set_xlabel("Temps (s)")
    axs[0, 0].set_ylabel("Accélération (m/s² ou g)")
    axs[0, 0].set_title("Perturbation en accélération")
    
    # --- Subplot 2: Spectre de déplacement relatif ---
    axs[0, 1].grid(True, which="both", ls="-")
    axs[0, 1].loglog(resultats['Frequences'], resultats['Deplacement'], color='purple', label='Amortissement 5%')
    axs[0, 1].set_xlabel("Fréquence de l'O.H (Hz)")
    axs[0, 1].set_ylabel("Déplacement (m)")
    axs[0, 1].set_title("Spectre de déplacement relatif")
    axs[0, 1].legend(loc='best')
    
    # Repérer et annoter le maximum pour le déplacement
    max_disp_idx = np.argmax(resultats['Deplacement'])
    max_disp_val = resultats['Deplacement'][max_disp_idx]
    max_disp_freq = resultats['Frequences'][max_disp_idx]
    axs[0, 1].annotate(f'Max: {max_disp_val:.2e} m\nFreq: {max_disp_freq:.2f} Hz',
                       xy=(max_disp_freq, max_disp_val),
                       xytext=(-40, -90),
                       textcoords='offset points',
                       arrowprops=dict(facecolor='black', shrink=0.05),
                       fontsize=10,
                       bbox=dict(boxstyle="round",pad=0.3, edgecolor="black", facecolor="white", alpha=0.5))

    # --- Subplot 3: Spectre de pseudo-vitesse ---
    axs[1, 0].grid(True, which="both", ls="-")
    axs[1, 0].loglog(resultats['Frequences'], resultats['Pseudo_Vitesse'], color='green', label='Amortissement 5%')
    axs[1, 0].set_xlabel("Fréquence de l'O.H (Hz)")
    axs[1, 0].set_ylabel("Pseudo-vitesse (m/s)")
    axs[1, 0].set_title("Spectre de pseudo-vitesse")
    axs[1, 0].legend(loc='best')
    
    # Repérer et annoter le maximum pour la vitesse
    max_vel_idx = np.argmax(resultats['Pseudo_Vitesse'])
    max_vel_val = resultats['Pseudo_Vitesse'][max_vel_idx]
    max_vel_freq = resultats['Frequences'][max_vel_idx]
    axs[1, 0].annotate(f'Max: {max_vel_val:.2e} m/s\nFreq: {max_vel_freq:.2f} Hz',
                       xy=(max_vel_freq, max_vel_val),
                       xytext=(-40, -90),
                       textcoords='offset points',
                       arrowprops=dict(facecolor='black', shrink=0.05),
                       fontsize=10,
                       bbox=dict(boxstyle="round",pad=0.3, edgecolor="black", facecolor="white", alpha=0.5))
    
    # --- Subplot 4: Spectre de pseudo-accélération ---
    axs[1, 1].grid(True, which="both", ls="-")
    axs[1, 1].loglog(resultats['Frequences'], resultats['Pseudo_Acceleration'], color='coral', label='Amortissement 5%')
    axs[1, 1].set_xlabel("Fréquence de l'O.H (Hz)")
    axs[1, 1].set_ylabel("Pseudo-accélération (m/s² ou g)")
    axs[1, 1].set_title("Spectre de pseudo-accélération")
    axs[1, 1].legend(loc='best')
    
    # Repérer et annoter le maximum pour l'accélération
    max_acc_idx = np.argmax(resultats['Pseudo_Acceleration'])
    max_acc_val = resultats['Pseudo_Acceleration'][max_acc_idx]
    max_acc_freq = resultats['Frequences'][max_acc_idx]
    axs[1, 1].annotate(f'Max: {max_acc_val:.2e} m/s²\nFreq: {max_acc_freq:.2f} Hz',
                       xy=(max_acc_freq, max_acc_val),
                       xytext=(-40,-90),
                       textcoords='offset points',
                       arrowprops=dict(facecolor='black', shrink=0.05),
                       fontsize=10,
                       bbox=dict(boxstyle="round", pad=0.3, edgecolor="black", facecolor="white", alpha=0.5))
    
    plt.tight_layout()
    plt.show()
    
    
# --- Exemple d'utilisation ---
"""
if __name__ == "__main__":

    dt = 0.01
    temps = np.arange(0, 10, dt)
    # Test avec un sinus dégradé
    accel = 1.0 * np.sin(2 * np.pi * 2.0 * temps) * np.exp(-5 * temps) * temps**5 #(m/s² ou g)
    
    #accel =  1.0 * np.sin(2.0 * temps + 3.0)
    
    # Paramètres de calcul
    amortissement = 0.05  # 5%
    frequences = np.logspace(np.log10(0.1), np.log10(50), 100) # De 0.1 à 50 Hz
    
    # Appel de la fonction
    print("Calcul en cours...")
    resultats = sro_nigam(frequences, accel, dt, amortissement)
    
    # Affichage puis export
    print("Calcul terminé.")
    print(f"SRO Max (Accélération): {np.max(resultats['Pseudo_Acceleration']):.4f}")
    
    affichage(temps, accel, resultats)
    

"""
    

# Lire le fichier Excel (feuille Calculs)
df_calculs = pd.read_excel("Spectre.xlsx", sheet_name="Calculs", header=None, engine="openpyxl")
dt = df_calculs.iloc[10, 3]  # Cellule D11 (index 0 donc ligne 10, colonne 3)
amort = df_calculs.iloc[19, 3]  # Cellule D20 (index 19, colonne 3)

# Lire l'accélérogramme (feuille A)
df_acc = pd.read_excel("Spectre.xlsx", sheet_name="A", engine="openpyxl")
accel_data = df_acc.iloc[2:, 1].values  # Colonne 2 (les données commencent à la ligne 3)
temps = df_acc.iloc[2:, 0].values  # Colonne 1 (sans la ligne de titre)

print(dt, amort, df_acc)
# Lancer le calcul (fonction supposée existante)
res = sro_nigam(frequencies, accel_data, dt, amort)  # Résultat attendu sous la forme d'une liste ou d'un tableau

# Affichage des données (si nécessaire pour vérification)
affichage(temps, accel_data, res)

# Charger le fichier existant avec openpyxl pour modification sans écrasement

try:
    # Charger le fichier Excel
    wb = load_workbook("Spectre.xlsx")
    
    # Créer/Réinitialiser la feuille "RES"
    if "RES" in wb.sheetnames:
        ws = wb["RES"]
        ws.delete_rows(1, ws.max_row)  # Vider les anciennes données
    else:
        ws = wb.create_sheet(title="RES")
    
    # Ajouter les en-têtes
    headers = ["Fréq. propres", "Déplacement rel.", "Ps. Vitesse", "Ps. Accélération"]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    
    # Ajouter les données
    for row_idx, row in enumerate(res.values(), start=1):  # Commence à la ligne 2
        for col_idx, value in enumerate(row, start=2):
            ws.cell(row=col_idx, column=row_idx, value=value)   #avoir 4 colonnes (feuille RES) à partir de 4 lignes (dict res)
    
    # Sauvegarder les modifications
    wb.save("Spectre.xlsx")
    wb.close() 
   
     
except Exception as e:
    print(f"Une erreur s'est produite : {e}")
